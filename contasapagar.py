import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, date
import os
from openpyxl import load_workbook

# Configuração da página
st.set_page_config(
    page_title="💼 Sistema Financeiro 2025",
    page_icon="💰",
    layout="wide",
    initial_sidebar_state="expanded"
)

VALID_USERS = {
    "Vinicius": "vinicius4223",
    "Flavio": "1234",
}

def check_login(username: str, password: str) -> bool:
    return VALID_USERS.get(username) == password

if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
    st.session_state.username = ""

if not st.session_state.logged_in:
    st.write("\n" * 5)
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.title("🔒 Login")
        username_input = st.text_input("Usuário:")
        password_input = st.text_input("Senha:", type="password")
        if st.button("Entrar"):
            if check_login(username_input, password_input):
                st.session_state.logged_in = True
                st.session_state.username = username_input
            else:
                st.error("Usuário ou senha inválidos.")
    st.stop()

# Constantes no início do arquivo (após as imports)
EXCEL_PAGAR = "Contas a pagar 2025.xlsx"
EXCEL_RECEBER = "Contas a receber 2025.xlsx"
ANEXOS_DIR = "anexos"  # Esta linha estava faltando
FULL_MONTHS = [f"{i:02d}" for i in range(1, 13)]

# Garante pastas de anexos (esta parte deve vir DEPOIS de definir ANEXOS_DIR)
for pasta in ["Contas a Pagar", "Contas a Receber"]:
    os.makedirs(os.path.join(ANEXOS_DIR, pasta), exist_ok=True)


st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600&display=swap');

html, body, [class*="css"] {
    font-family: 'Inter', sans-serif;
    line-height: 1.6;
}

/* ----------------- Cabeçalho ----------------- */
.header {
    background: linear-gradient(135deg, #4a00e0, #8e2de2);
    color: white;
    padding: 2rem;
    border-radius: 16px;
    box-shadow: 0 8px 24px rgba(0, 0, 0, 0.2);
    margin-bottom: 2rem;
    text-align: center;
}

/* ----------------- Cards de Métricas ----------------- */
.metric-card {
    border-radius: 16px;
    padding: 1.5rem;
    border-left: 5px solid #8e2de2;
    transition: transform 0.3s ease, box-shadow 0.3s ease;
    box-shadow: 0 3px 10px rgba(0, 0, 0, 0.05);
    margin-bottom: 1rem;
}

.metric-card:hover {
    transform: translateY(-5px);
    box-shadow: 0 10px 20px rgba(0, 0, 0, 0.1);
}

.metric-value {
    font-size: 2.5rem;
    font-weight: 600;
    margin-bottom: 0.25rem;
}

.metric-label {
    font-size: 0.85rem;
    text-transform: uppercase;
    letter-spacing: 1px;
}

/* Tema Claro */
.stApp:not([data-theme="dark"]) .metric-card {
    background: #f9fafe;
}
.stApp:not([data-theme="dark"]) .metric-value {
    color: #2c3e50;
}
.stApp:not([data-theme="dark"]) .metric-label {
    color: #7f8c8d;
}

/* Tema Escuro */
.stApp[data-theme="dark"] .metric-card {
    background: #23242b;
}
.stApp[data-theme="dark"] .metric-value {
    color: #ffffff;
}
.stApp[data-theme="dark"] .metric-label {
    color: #b0b3bd;
}

/* ----------------- Tabs ----------------- */
.stTabs [role="tablist"] {
    gap: 10px;
    margin-bottom: 1rem;
}
.stTabs [role="tab"] {
    padding: 10px 20px;
    border-radius: 10px 10px 0 0;
    font-weight: 500;
    transition: all 0.3s ease-in-out;
    border: none;
}
.stApp:not([data-theme="dark"]) .stTabs [role="tab"] {
    background: #e0e3ec;
    color: #34495e;
}
.stApp[data-theme="dark"] .stTabs [role="tab"] {
    background: #2b2d3a;
    color: #bbb;
}
.stTabs [role="tab"][aria-selected="true"] {
    background: #8e2de2 !important;
    color: white !important;
}

/* ----------------- Gráfico Container ----------------- */
.chart-container {
    border-radius: 16px;
    padding: 1.5rem;
    margin-bottom: 2rem;
    transition: box-shadow 0.3s ease;
}
.stApp:not([data-theme="dark"]) .chart-container {
    background: #ffffff;
    box-shadow: 0 2px 8px rgba(0, 0, 0, 0.05);
}
.stApp[data-theme="dark"] .chart-container {
    background: #1e1f26;
    box-shadow: 0 2px 12px rgba(0, 0, 0, 0.2);
}

/* ----------------- Botões ----------------- */
.stButton > button,
.stDownloadButton > button {
    border-radius: 10px;
    padding: 10px 20px;
    font-weight: 500;
    transition: all 0.3s ease;
    border: none;
    box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
    cursor: pointer;
}
.stApp:not([data-theme="dark"]) .stButton > button,
.stApp:not([data-theme="dark"]) .stDownloadButton > button {
    background-color: #4a00e0;
    color: white;
}
.stApp[data-theme="dark"] .stButton > button,
.stApp[data-theme="dark"] .stDownloadButton > button {
    background-color: #9c6df2;
    color: white;
}
.stButton > button:hover,
.stDownloadButton > button:hover {
    filter: brightness(0.95);
    transform: scale(1.02);
}

/* ----------------- Inputs ----------------- */
input, select, textarea, .stTextInput > div > input {
    border-radius: 8px !important;
    padding: 0.5rem !important;
    font-size: 0.95rem !important;
    border: 1px solid #ccc !important;
}

.stApp[data-theme="dark"] input,
.stApp[data-theme="dark"] select,
.stApp[data-theme="dark"] textarea {
    background-color: #2b2d3a !important;
    color: white !important;
    border: 1px solid #444 !important;
}

/* ----------------- File Uploader ----------------- */
section[data-testid="stFileUploader"] div[role="button"] {
    border-radius: 8px;
    padding: 0.75rem 1rem;
    background-color: #8e2de2;
    color: white;
    border: none;
    font-weight: 500;
}
section[data-testid="stFileUploader"] div[role="button"]:hover {
    background-color: #762bd0;
}

/* ----------------- DataFrame ----------------- */
.css-1d391kg, .stDataFrame, .stTable {
    border-radius: 10px !important;
    overflow: hidden;
}
thead tr th {
    background-color: #8e2de2 !important;
    color: white !important;
    font-weight: 600;
    padding: 0.5rem;
}
tbody tr td {
    padding: 0.4rem;
}
.stApp[data-theme="dark"] .stDataFrame tbody tr td {
    background-color: #1f2028 !important;
    color: #e0e0e0 !important;
}
.stApp:not([data-theme="dark"]) .stDataFrame tbody tr td {
    background-color: #ffffff !important;
    color: #2c3e50 !important;
}
</style>
""", unsafe_allow_html=True)


def get_existing_sheets(excel_path: str) -> list[str]:
    try:
        wb = pd.ExcelFile(excel_path)
        numeric_sheets = []
        for s in wb.sheet_names:
            nome = s.strip()
            if nome.lower() == "tutorial":
                continue
            if nome.isdigit():
                nome_formatado = f"{int(nome):02d}"
                numeric_sheets.append(nome_formatado)
        return sorted(set(numeric_sheets))
    except Exception as e:
        st.error(f"Erro ao ler abas do arquivo: {e}")
        return []

def load_data(excel_path: str, sheet_name: str) -> pd.DataFrame:
    cols = [
        "data_nf", "forma_pagamento", "fornecedor", "os",
        "vencimento", "valor", "estado", "situacao", "boleto", "comprovante"
    ]
    
    if not os.path.isfile(excel_path):
        return pd.DataFrame(columns=cols + ["status_pagamento"])

    try:
        # Mapeia abas numéricas ("04" → "4")
        sheet_lookup = {}
        with pd.ExcelFile(excel_path) as wb:
            for s in wb.sheet_names:
                nome = s.strip()
                if nome.lower() != "tutorial" and nome.isdigit():
                    sheet_lookup[f"{int(nome):02d}"] = nome
        
        if sheet_name not in sheet_lookup:
            return pd.DataFrame(columns=cols + ["status_pagamento"])
            
        real_sheet = sheet_lookup[sheet_name]
        df = pd.read_excel(excel_path, sheet_name=real_sheet, skiprows=7, header=0)

        # Renomeia colunas
        rename_map = {}
        for col in df.columns:
            nome = str(col).strip().lower()
            if ("data" in nome and "nf" in nome) or "data da nota fiscal" in nome:
                rename_map[col] = "data_nf"
            elif "forma" in nome and "pagamento" in nome:
                rename_map[col] = "forma_pagamento"
            elif nome == "descrição":
                rename_map[col] = "forma_pagamento"
            elif nome == "fornecedor" or "cliente" in nome:
                rename_map[col] = "fornecedor"
            elif "os" in nome or nome == "documento":
                rename_map[col] = "os"
            elif "vencimento" in nome:
                rename_map[col] = "vencimento"
            elif "valor" in nome:
                rename_map[col] = "valor"
            elif nome == "estado":
                rename_map[col] = "estado"
            elif "situa" in nome:
                rename_map[col] = "situacao"
            elif "comprov" in nome:
                rename_map[col] = "comprovante"
            elif "boleto" in nome:
                rename_map[col] = "boleto"

        df = df.rename(columns=rename_map)
        df = df[[c for c in df.columns if c in cols]]

        # Garante colunas mínimas
        for obrig in ["fornecedor", "valor"]:
            if obrig not in df.columns:
                df[obrig] = pd.NA

        df = df.dropna(subset=["fornecedor", "valor"], how="all").reset_index(drop=True)

        # Converte tipos
        df["vencimento"] = pd.to_datetime(df["vencimento"], errors="coerce")
        df["valor"] = pd.to_numeric(df["valor"], errors="coerce")

        # Detecta modo: Pagar ou Receber
        is_receber = (excel_path == EXCEL_RECEBER)

        # Monta status_pagamento
        status_list = []
        hoje = datetime.now().date()
        for _, row in df.iterrows():
            estado_atual = str(row.get("estado", "")).strip().lower()
            if estado_atual == ("recebido" if is_receber else "pago"):
                status_list.append("Recebido" if is_receber else "Pago")
            else:
                data_venc = row["vencimento"].date() if pd.notna(row["vencimento"]) else None
                if data_venc:
                    if data_venc < hoje:
                        status_list.append("Em Atraso")
                    else:
                        status_list.append("A Receber" if is_receber else "Em Aberto")
                else:
                    status_list.append("Sem Data")

        df["status_pagamento"] = status_list
        return df

    except Exception as e:
        st.error(f"Erro ao carregar dados: {e}")
        return pd.DataFrame(columns=cols + ["status_pagamento"])

def save_data(excel_path: str, sheet_name: str, df: pd.DataFrame) -> bool:
    try:
        wb = load_workbook(excel_path)
        
        if sheet_name not in wb.sheetnames:
            st.error(f"A aba '{sheet_name}' não existe no arquivo.")
            return False
            
        ws = wb[sheet_name]
        header_row = 8
        
        headers = [
            str(ws.cell(row=header_row, column=col).value).strip().lower()
            for col in range(2, ws.max_column + 1)
        ]

        field_map = {
            "data_nf": ["data documento", "data_nf", "data n/f", "data da nota fiscal"],
            "forma_pagamento": ["descrição", "forma_pagamento", "forma de pagamento"],
            "fornecedor": ["fornecedor"],
            "os": ["documento", "os", "os interna"],
            "vencimento": ["vencimento"],
            "valor": ["valor"],
            "estado": ["estado"],
            "boleto": ["boleto", "boleto anexo"],
            "comprovante": ["comprovante", "comprovante de pagto"]
        }

        col_pos = {}
        for key, names in field_map.items():
            idx = next((i for i, h in enumerate(headers) if h in names), None)
            col_pos[key] = idx + 2 if idx is not None else None

        for i, row in df.iterrows():
            excel_row = header_row + 1 + i
            for key, col in col_pos.items():
                if not col or key == "situacao":
                    continue
                    
                val = row.get(key, "")
                if key in ("data_nf", "vencimento"):
                    try:
                        val = pd.to_datetime(val, errors="coerce")
                        if pd.notna(val):
                            val = val.to_pydatetime()
                        else:
                            continue
                    except:
                        continue
                elif key == "valor":
                    try:
                        val = float(val)
                    except:
                        val = None

                ws.cell(row=excel_row, column=col, value=val)

        wb.save(excel_path)
        return True
        
    except Exception as e:
        st.error(f"Erro ao salvar dados: {e}")
        return False
        
def add_record(excel_path: str, sheet_name: str, record: dict) -> bool:
    try:
        wb = load_workbook(excel_path)
        
        if sheet_name not in wb.sheetnames:
            numeric = [s for s in wb.sheetnames if s.isdigit()]
            template_ws = wb[numeric[0]] if numeric else wb[wb.sheetnames[0]]
            ws = wb.copy_worksheet(template_ws)
            ws.title = sheet_name
        else:
            ws = wb[sheet_name]

        header_row = 8
        headers = [
            str(ws.cell(row=header_row, column=col).value).strip().lower()
            for col in range(2, ws.max_column + 1)
        ]

        field_map = {
            "data_nf": ["data documento", "data_nf", "data n/f", "data da nota fiscal"],
            "forma_pagamento": ["descrição", "forma_pagamento", "forma de pagamento"],
            "fornecedor": ["fornecedor"],
            "os": ["documento", "os", "os interna"],
            "vencimento": ["vencimento"],
            "valor": ["valor"],
            "estado": ["estado"],
            "boleto": ["boleto", "boleto anexo"],
            "comprovante": ["comprovante", "comprovante de pagto"]
        }

        col_pos = {}
        for key, names in field_map.items():
            idx = next((i for i, h in enumerate(headers) if h in names), None)
            col_pos[key] = idx + 2 if idx is not None else None

        col_forn = col_pos.get("fornecedor", 2)

        # ✅ CORRIGIDO: encontra próxima linha vazia com base no fornecedor
        next_row = header_row + 1
        while ws.cell(row=next_row, column=col_forn).value:
            next_row += 1

        for key, col in col_pos.items():
            if not col or key == "situacao":
                continue

            val = record.get(key, "")
            if key in ("data_nf", "vencimento"):
                try:
                    dt = pd.to_datetime(val, errors="coerce")
                    if pd.notna(dt):
                        val = dt.to_pydatetime()
                    else:
                        continue
                except:
                    continue
            elif key == "valor":
                try:
                    val = float(val)
                except:
                    val = None

            ws.cell(row=next_row, column=col, value=val)

        wb.save(excel_path)
        return True

    except Exception as e:
        st.error(f"Erro ao adicionar registro: {e}")
        return False

st.markdown("""
<div style="text-align: center; color: #4B8BBE; margin-bottom: 10px;">
    <h1>💼 Sistema Financeiro 2025</h1>
    <p style="color: #555; font-size: 16px;">Dashboard avançado com estatísticas e gráficos interativos.</p>
</div>
""", unsafe_allow_html=True)
st.markdown("---")

# 👤 Mostra usuário logado
st.sidebar.markdown(f"**Logado:** {st.session_state.username}")

# 🔘 NAVEGAÇÃO
page = st.sidebar.radio("Ir para:", ["Dashboard", "Contas a Pagar", "Contas a Receber"])

# Dashboard Modernizado
if page == "Dashboard":
    if not os.path.isfile(EXCEL_PAGAR):
        st.error(f"Arquivo '{EXCEL_PAGAR}' não encontrado. Verifique o caminho.")
    if not os.path.isfile(EXCEL_RECEBER):
        st.error(f"Arquivo '{EXCEL_RECEBER}' não encontrado. Verifique o caminho.")
    if not os.path.isfile(EXCEL_PAGAR) or not os.path.isfile(EXCEL_RECEBER):
        st.stop()
    
    sheets_p = get_existing_sheets(EXCEL_PAGAR)
    sheets_r = get_existing_sheets(EXCEL_RECEBER)
    
    # Layout com tabs modernas
    tab1, tab2 = st.tabs(["📥 Contas a Pagar", "📤 Contas a Receber"])
    
    with tab1:
        if not sheets_p:
            st.warning("Nenhuma aba válida encontrada em Contas a Pagar")
        else:
            df_all_p = pd.concat([load_data(EXCEL_PAGAR, s) for s in sheets_p], ignore_index=True)
            
            if df_all_p.empty:
                st.info("Nenhum dado encontrado nas planilhas de Contas a Pagar")
            else:
                # Métricas principais
                total_p = df_all_p["valor"].sum()
                num_lanc_p = len(df_all_p)
                media_p = df_all_p["valor"].mean() if num_lanc_p else 0
                atrasados_p = df_all_p[df_all_p["status_pagamento"] == "Em Atraso"]
                num_atras_p = len(atrasados_p)
                perc_atras_p = (num_atras_p / num_lanc_p * 100) if num_lanc_p else 0
                
                # Layout de métricas
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-label">Total a Pagar</div>
                        <div class="metric-value">R$ {total_p:,.2f}</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with col2:
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-label">Lançamentos</div>
                        <div class="metric-value">{num_lanc_p}</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with col3:
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-label">Média por Conta</div>
                        <div class="metric-value">R$ {media_p:,.2f}</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with col4:
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-label">Em Atraso</div>
                        <div class="metric-value">{perc_atras_p:.1f}%</div>
                        <div style="font-size: 0.8rem; color: {'#e74c3c' if perc_atras_p > 10 else '#27ae60'}">
                            ({num_atras_p} conta{'s' if num_atras_p != 1 else ''})
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                
                st.markdown("---")
                
                # Gráfico de distribuição por status
                st.markdown("#### 📊 Distribuição por Status")
                status_counts_p = (
                    df_all_p["status_pagamento"]
                    .value_counts()
                    .rename_axis("status")
                    .reset_index(name="contagem")
                )
                
                fig_status = px.pie(
                    status_counts_p,
                    values="contagem",
                    names="status",
                    hole=0.4,
                    color_discrete_sequence=px.colors.qualitative.Pastel
                )
                fig_status.update_traces(
                    textposition="inside",
                    textinfo="percent+label",
                    hovertemplate="<b>%{label}</b><br>%{value} contas (%{percent})"
                )
                fig_status.update_layout(
                    showlegend=False,
                    margin=dict(l=20, r=20, t=30, b=20),
                    height=350
                )
                
                col1, col2 = st.columns([3, 1])
                with col1:
                    st.plotly_chart(fig_status, use_container_width=True)
                
                with col2:
                    st.markdown("""
                    <div style="background: #f8f9fa; padding: 1rem; border-radius: 10px;">
                        <h4 style="margin-top: 0;">Legenda</h4>
                        <div style="display: flex; align-items: center; margin-bottom: 8px;">
                            <div style="width: 12px; height: 12px; background: #636EFA; border-radius: 50%; margin-right: 8px;"></div>
                            <span>Em Aberto</span>
                        </div>
                        <div style="display: flex; align-items: center; margin-bottom: 8px;">
                            <div style="width: 12px; height: 12px; background: #EF553B; border-radius: 50%; margin-right: 8px;"></div>
                            <span>Em Atraso</span>
                        </div>
                        <div style="display: flex; align-items: center; margin-bottom: 8px;">
                            <div style="width: 12px; height: 12px; background: #00CC96; border-radius: 50%; margin-right: 8px;"></div>
                            <span>Pago</span>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                
                st.markdown("---")
                
                # Evolução mensal
                st.markdown("#### 📈 Evolução Mensal")
                df_all_p["mes_ano"] = df_all_p["vencimento"].dt.to_period("M")
                monthly_group_p = (
                    df_all_p
                    .groupby("mes_ano")
                    .agg(
                        total_mes=("valor", "sum"),
                        pagos_mes=("valor", lambda x: x[df_all_p.loc[x.index, "status_pagamento"] == "Pago"].sum()),
                        pendentes_mes=("valor", lambda x: x[df_all_p.loc[x.index, "status_pagamento"] != "Pago"].sum())
                    )
                    .reset_index()
                )
                monthly_group_p["mes_ano_str"] = monthly_group_p["mes_ano"].dt.strftime("%b/%Y")
                
                fig_evolucao = go.Figure()
                fig_evolucao.add_trace(go.Scatter(
                    x=monthly_group_p["mes_ano_str"],
                    y=monthly_group_p["total_mes"],
                    name="Total",
                    line=dict(color="#6e8efb", width=3),
                    mode="lines+markers",
                    hovertemplate="<b>%{x}</b><br>Total: R$ %{y:,.2f}<extra></extra>"
                ))
                fig_evolucao.add_trace(go.Scatter(
                    x=monthly_group_p["mes_ano_str"],
                    y=monthly_group_p["pagos_mes"],
                    name="Pagas",
                    line=dict(color="#00CC96", width=2),
                    mode="lines+markers",
                    hovertemplate="<b>%{x}</b><br>Pagas: R$ %{y:,.2f}<extra></extra>"
                ))
                fig_evolucao.add_trace(go.Scatter(
                    x=monthly_group_p["mes_ano_str"],
                    y=monthly_group_p["pendentes_mes"],
                    name="Pendentes",
                    line=dict(color="#EF553B", width=2),
                    mode="lines+markers",
                    hovertemplate="<b>%{x}</b><br>Pendentes: R$ %{y:,.2f}<extra></extra>"
                ))
                
                fig_evolucao.update_layout(
                    hovermode="x unified",
                    legend=dict(
                        orientation="h",
                        yanchor="bottom",
                        y=1.02,
                        xanchor="right",
                        x=1
                    ),
                    margin=dict(l=20, r=20, t=30, b=20),
                    height=400,
                    xaxis_title="Mês/Ano",
                    yaxis_title="Valor (R$)",
                    plot_bgcolor="rgba(0,0,0,0)",
                    paper_bgcolor="rgba(0,0,0,0)"
                )
                
                st.plotly_chart(fig_evolucao, use_container_width=True)
                
                # Top 10 fornecedores
                st.markdown("---")
                st.markdown("#### 🏆 Top 10 Fornecedores")
                top_fornecedores = (
                    df_all_p.groupby("fornecedor")
                    .agg(total=("valor", "sum"), contagem=("valor", "count"))
                    .sort_values("total", ascending=False)
                    .head(10)
                    .reset_index()
                )
                
                fig_fornecedores = px.bar(
                    top_fornecedores,
                    x="total",
                    y="fornecedor",
                    orientation="h",
                    color="contagem",
                    color_continuous_scale="Blues",
                    labels={"total": "Valor Total (R$)", "fornecedor": "", "contagem": "Nº Contas"},
                    hover_data={"contagem": True}
                )
                fig_fornecedores.update_layout(
                    height=500,
                    xaxis_title="Valor Total (R$)",
                    yaxis_title="",
                    yaxis={"categoryorder": "total ascending"},
                    margin=dict(l=20, r=20, t=30, b=20),
                    coloraxis_colorbar=dict(title="Nº Contas")
                )
                
                st.plotly_chart(fig_fornecedores, use_container_width=True)
                
                # Download dos dados
                st.markdown("---")
                with st.expander("💾 Exportar Dados", expanded=False):
                    try:
                        with open(EXCEL_PAGAR, "rb") as f:
                            st.download_button(
                                label="Baixar Planilha Completa (Contas a Pagar)",
                                data=f.read(),
                                file_name=EXCEL_PAGAR,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                    except Exception as e:
                        st.error(f"Erro ao preparar download: {e}")

    with tab2:
        if not sheets_r:
            st.warning("Nenhuma aba válida encontrada em Contas a Receber")
        else:
            df_all_r = pd.concat([load_data(EXCEL_RECEBER, s) for s in sheets_r], ignore_index=True)
            
            if df_all_r.empty:
                st.info("Nenhum dado encontrado nas planilhas de Contas a Receber")
            else:
                # Métricas principais
                total_r = df_all_r["valor"].sum()
                num_lanc_r = len(df_all_r)
                media_r = df_all_r["valor"].mean() if num_lanc_r else 0
                atrasados_r = df_all_r[df_all_r["status_pagamento"] == "Em Atraso"]
                num_atras_r = len(atrasados_r)
                perc_atras_r = (num_atras_r / num_lanc_r * 100) if num_lanc_r else 0
                
                # Layout de métricas
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-label">Total a Receber</div>
                        <div class="metric-value">R$ {total_r:,.2f}</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with col2:
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-label">Lançamentos</div>
                        <div class="metric-value">{num_lanc_r}</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with col3:
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-label">Média por Conta</div>
                        <div class="metric-value">R$ {media_r:,.2f}</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with col4:
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-label">Em Atraso</div>
                        <div class="metric-value">{perc_atras_r:.1f}%</div>
                        <div style="font-size: 0.8rem; color: {'#e74c3c' if perc_atras_r > 10 else '#27ae60'}">
                            ({num_atras_r} conta{'s' if num_atras_r != 1 else ''})
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                
                st.markdown("---")
                
                # Gráfico de distribuição por status
                st.markdown("#### 📊 Distribuição por Status")
                status_counts_r = (
                    df_all_r["status_pagamento"]
                    .value_counts()
                    .rename_axis("status")
                    .reset_index(name="contagem")
                )
                
                fig_status = px.pie(
                    status_counts_r,
                    values="contagem",
                    names="status",
                    hole=0.4,
                    color_discrete_sequence=px.colors.qualitative.Pastel
                )
                fig_status.update_traces(
                    textposition="inside",
                    textinfo="percent+label",
                    hovertemplate="<b>%{label}</b><br>%{value} contas (%{percent})"
                )
                fig_status.update_layout(
                    showlegend=False,
                    margin=dict(l=20, r=20, t=30, b=20),
                    height=350
                )
                
                col1, col2 = st.columns([3, 1])
                with col1:
                    st.plotly_chart(fig_status, use_container_width=True)
                
                with col2:
                    st.markdown("""
                    <div style="background: #f8f9fa; padding: 1rem; border-radius: 10px;">
                        <h4 style="margin-top: 0;">Legenda</h4>
                        <div style="display: flex; align-items: center; margin-bottom: 8px;">
                            <div style="width: 12px; height: 12px; background: #636EFA; border-radius: 50%; margin-right: 8px;"></div>
                            <span>A Receber</span>
                        </div>
                        <div style="display: flex; align-items: center; margin-bottom: 8px;">
                            <div style="width: 12px; height: 12px; background: #EF553B; border-radius: 50%; margin-right: 8px;"></div>
                            <span>Em Atraso</span>
                        </div>
                        <div style="display: flex; align-items: center; margin-bottom: 8px;">
                            <div style="width: 12px; height: 12px; background: #00CC96; border-radius: 50%; margin-right: 8px;"></div>
                            <span>Recebido</span>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                
                st.markdown("---")
                
                # Evolução mensal
                st.markdown("#### 📈 Evolução Mensal")
                df_all_r["mes_ano"] = df_all_r["vencimento"].dt.to_period("M")
                monthly_group_r = (
                    df_all_r
                    .groupby("mes_ano")
                    .agg(
                        total_mes=("valor", "sum"),
                        recebidos_mes=("valor", lambda x: x[df_all_r.loc[x.index, "status_pagamento"] == "Recebido"].sum()),
                        pendentes_mes=("valor", lambda x: x[df_all_r.loc[x.index, "status_pagamento"] != "Recebido"].sum())
                    )
                    .reset_index()
                )
                monthly_group_r["mes_ano_str"] = monthly_group_r["mes_ano"].dt.strftime("%b/%Y")
                
                fig_evolucao = go.Figure()
                fig_evolucao.add_trace(go.Scatter(
                    x=monthly_group_r["mes_ano_str"],
                    y=monthly_group_r["total_mes"],
                    name="Total",
                    line=dict(color="#6e8efb", width=3),
                    mode="lines+markers",
                    hovertemplate="<b>%{x}</b><br>Total: R$ %{y:,.2f}<extra></extra>"
                ))
                fig_evolucao.add_trace(go.Scatter(
                    x=monthly_group_r["mes_ano_str"],
                    y=monthly_group_r["recebidos_mes"],
                    name="Recebidos",
                    line=dict(color="#00CC96", width=2),
                    mode="lines+markers",
                    hovertemplate="<b>%{x}</b><br>Recebidos: R$ %{y:,.2f}<extra></extra>"
                ))
                fig_evolucao.add_trace(go.Scatter(
                    x=monthly_group_r["mes_ano_str"],
                    y=monthly_group_r["pendentes_mes"],
                    name="Pendentes",
                    line=dict(color="#EF553B", width=2),
                    mode="lines+markers",
                    hovertemplate="<b>%{x}</b><br>Pendentes: R$ %{y:,.2f}<extra></extra>"
                ))
                
                fig_evolucao.update_layout(
                    hovermode="x unified",
                    legend=dict(
                        orientation="h",
                        yanchor="bottom",
                        y=1.02,
                        xanchor="right",
                        x=1
                    ),
                    margin=dict(l=20, r=20, t=30, b=20),
                    height=400,
                    xaxis_title="Mês/Ano",
                    yaxis_title="Valor (R$)",
                    plot_bgcolor="rgba(0,0,0,0)",
                    paper_bgcolor="rgba(0,0,0,0)"
                )
                
                st.plotly_chart(fig_evolucao, use_container_width=True)
                
                # Top 10 clientes
                st.markdown("---")
                st.markdown("#### 🏆 Top 10 Clientes")
                top_clientes = (
                    df_all_r.groupby("fornecedor")
                    .agg(total=("valor", "sum"), contagem=("valor", "count"))
                    .sort_values("total", ascending=False)
                    .head(10)
                    .reset_index()
                )
                
                fig_clientes = px.bar(
                    top_clientes,
                    x="total",
                    y="fornecedor",
                    orientation="h",
                    color="contagem",
                    color_continuous_scale="Blues",
                    labels={"total": "Valor Total (R$)", "fornecedor": "", "contagem": "Nº Contas"},
                    hover_data={"contagem": True}
                )
                fig_clientes.update_layout(
                    height=500,
                    xaxis_title="Valor Total (R$)",
                    yaxis_title="",
                    yaxis={"categoryorder": "total ascending"},
                    margin=dict(l=20, r=20, t=30, b=20),
                    coloraxis_colorbar=dict(title="Nº Contas")
                )
                
                st.plotly_chart(fig_clientes, use_container_width=True)
                
                # Download dos dados
                st.markdown("---")
                with st.expander("💾 Exportar Dados", expanded=False):
                    try:
                        with open(EXCEL_RECEBER, "rb") as f:
                            st.download_button(
                                label="Baixar Planilha Completa (Contas a Receber)",
                                data=f.read(),
                                file_name=EXCEL_RECEBER,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                    except Exception as e:
                        st.error(f"Erro ao preparar download: {e}")


elif page == "Contas a Pagar":
    st.subheader("🗂️ Contas a Pagar")
    
    # Verificação do arquivo
    if not os.path.isfile(EXCEL_PAGAR):
        st.error(f"Arquivo '{EXCEL_PAGAR}' não encontrado. Verifique o caminho.")
        st.stop()
    
    # Seleção do mês
    existing_sheets = get_existing_sheets(EXCEL_PAGAR)
    aba = st.selectbox("Selecione o mês:", FULL_MONTHS, index=FULL_MONTHS.index(date.today().strftime("%B")))
    
    # Carrega os dados
    df = load_data(EXCEL_PAGAR, aba)
    
    # Adiciona lançamentos temporários se existirem
    if "lista_lancamentos" in st.session_state and st.session_state.lista_lancamentos:
        df_temp = pd.DataFrame(st.session_state.lista_lancamentos)
        df = pd.concat([df, df_temp], ignore_index=True)
    
    # Filtro principal
    view_sel = st.radio("Visualizar:", ["Todos", "Pagas", "Pendentes"], horizontal=True)
    
    if view_sel == "Pagas":
        df_display = df[df["status_pagamento"] == "Pago"].copy()
    elif view_sel == "Pendentes":
        df_display = df[df["status_pagamento"] != "Pago"].copy()
    else:
        df_display = df.copy()
    
    # Adiciona numeração das linhas
    df_display.insert(0, '#', range(1, len(df_display) + 1))
    
    # Filtros avançados
    with st.expander("🔍 Filtros Avançados", expanded=False):
        col1, col2 = st.columns(2)
        with col1:
            fornecedor_filtro = st.selectbox(
                "Fornecedor",
                ["Todos"] + sorted(df["fornecedor"].dropna().unique().tolist()))
        with col2:
            status_filtro = st.selectbox(
                "Status",
                ["Todos"] + sorted(df["status_pagamento"].dropna().unique().tolist()))
    
    # Aplica filtros
    if fornecedor_filtro != "Todos":
        df_display = df_display[df_display["fornecedor"] == fornecedor_filtro]
    if status_filtro != "Todos":
        df_display = df_display[df_display["status_pagamento"] == status_filtro]
    
    # Exibe a tabela principal
    st.markdown("### 📋 Lançamentos")
    if df_display.empty:
        st.warning("Nenhum registro encontrado com os filtros selecionados.")
    else:
        # Seleciona e formata colunas para exibição
        cols_padrao = ['#', 'data_nf', 'fornecedor', 'valor', 'vencimento', 'status_pagamento', 'estado']
        cols_disponiveis = [c for c in cols_padrao if c in df_display.columns]
        
        df_exibicao = df_display[cols_disponiveis].copy()
        
        # Formatação
        if 'valor' in df_exibicao.columns:
            df_exibicao['valor'] = df_exibicao['valor'].apply(lambda x: f"R$ {float(x):,.2f}")
        if 'vencimento' in df_exibicao.columns:
            df_exibicao['vencimento'] = pd.to_datetime(df_exibicao['vencimento']).dt.strftime('%d/%m/%Y')
        if 'data_nf' in df_exibicao.columns:
            df_exibicao['data_nf'] = pd.to_datetime(df_exibicao['data_nf']).dt.strftime('%d/%m/%Y')
        
        st.dataframe(df_exibicao, height=400, use_container_width=True)
    
    # Seção de Edição
    with st.expander("✏️ Editar Registro", expanded=False):
        if not df_display.empty:
            idx_edicao = st.number_input(
                "Número da linha para editar:",
                min_value=1,
                max_value=len(df_display),
                step=1,
                key="edit_idx_pagar"
            )
            
            registro = df_display[df_display['#'] == idx_edicao].iloc[0]
            original_idx = df[df['fornecedor'] == registro['fornecedor']].index[0]
            
            col1, col2 = st.columns(2)
            with col1:
                novo_valor = st.number_input(
                    "Valor (R$):",
                    value=float(registro['valor']),
                    step=0.01,
                    key="edit_valor_pagar"
                )
                novo_vencimento = st.date_input(
                    "Vencimento:",
                    value=pd.to_datetime(registro['vencimento']).date(),
                    key="edit_venc_pagar"
                )
            with col2:
                novo_estado = st.selectbox(
                    "Estado:",
                    options=["Em Aberto", "Pago"],
                    index=0 if registro['estado'] == "Em Aberto" else 1,
                    key="edit_estado_pagar"
                )
                nova_situacao = st.selectbox(
                    "Situação:",
                    options=["Em Atraso", "Pago", "Em Aberto"],
                    index=0 if registro['situacao'] == "Em Atraso" else 1 if registro['situacao'] == "Pago" else 2,
                    key="edit_situacao_pagar"
                )
            
            if st.button("💾 Salvar Alterações", key="save_edit_pagar"):
                df.at[original_idx, 'valor'] = novo_valor
                df.at[original_idx, 'vencimento'] = novo_vencimento
                df.at[original_idx, 'estado'] = novo_estado
                df.at[original_idx, 'situacao'] = nova_situacao
                
                if save_data(EXCEL_PAGAR, aba, df):
                    st.success("Registro atualizado com sucesso!")
                    st.experimental_rerun()
                else:
                    st.error("Erro ao salvar alterações.")
    
    # Seção para Remover Registros
    with st.expander("🗑️ Remover Registro", expanded=False):
        if "lista_lancamentos" in st.session_state and st.session_state.lista_lancamentos:
            df_temp = pd.DataFrame(st.session_state.lista_lancamentos)
            df_temp.insert(0, '#', range(1, len(df_temp) + 1))
            
            st.dataframe(df_temp, height=150)
            
            idx_remocao = st.number_input(
                "Número da linha para remover:",
                min_value=1,
                max_value=len(df_temp),
                step=1,
                key="remove_idx_pagar"
            )
            
            if st.button("Remover Registro", key="btn_remove_pagar"):
                try:
                    st.session_state.lista_lancamentos.pop(idx_remocao - 1)
                    st.success("Registro removido com sucesso!")
                    st.experimental_rerun()
                except Exception as e:
                    st.error(f"Erro ao remover registro: {e}")
    
    # Seção para Adicionar Novos Registros
    with st.expander("➕ Adicionar Nova Conta", expanded=False):
        col1, col2 = st.columns(2)
        with col1:
            nova_data_nf = st.date_input("Data N/F:", value=date.today())
            nova_descricao = st.text_input("Descrição:")
            novo_fornecedor = st.text_input("Fornecedor:")
        with col2:
            novo_os = st.text_input("Documento/OS:")
            novo_vencimento = st.date_input("Vencimento:", value=date.today())
            novo_valor = st.number_input("Valor (R$):", min_value=0.01, step=0.01)
        
        novo_estado = st.selectbox("Estado:", ["Em Aberto", "Pago"])
        nova_situacao = st.selectbox("Situação:", ["Em Atraso", "Pago", "Em Aberto"])
        
        # Upload de arquivos
        col_anexo1, col_anexo2 = st.columns(2)
        with col_anexo1:
            boleto_file = st.file_uploader("Boleto (opcional):", type=["pdf", "jpg", "png"])
        with col_anexo2:
            comprovante_file = st.file_uploader("Comprovante (opcional):", type=["pdf", "jpg", "png"])
        
        if st.button("Adicionar Conta", key="btn_add_pagar"):
            novo_registro = {
                "data_nf": nova_data_nf,
                "forma_pagamento": nova_descricao,
                "fornecedor": novo_fornecedor,
                "os": novo_os,
                "vencimento": novo_vencimento,
                "valor": novo_valor,
                "estado": novo_estado,
                "situacao": nova_situacao,
                "status_pagamento": "Pago" if novo_estado == "Pago" else "Pendente"
            }
            
            # Processa anexos
            if boleto_file:
                boleto_path = os.path.join(ANEXOS_DIR, "Contas a Pagar", f"boleto_{uuid.uuid4()}.{boleto_file.name.split('.')[-1]}")
                with open(boleto_path, "wb") as f:
                    f.write(boleto_file.getbuffer())
                novo_registro["boleto"] = boleto_path
            
            if comprovante_file:
                comprovante_path = os.path.join(ANEXOS_DIR, "Contas a Pagar", f"comprovante_{uuid.uuid4()}.{comprovante_file.name.split('.')[-1]}")
                with open(comprovante_path, "wb") as f:
                    f.write(comprovante_file.getbuffer())
                novo_registro["comprovante"] = comprovante_path
            
            if add_record(EXCEL_PAGAR, aba, novo_registro):
                st.success("Conta adicionada com sucesso!")
                st.experimental_rerun()
            else:
                st.error("Erro ao adicionar nova conta.")
    
    # Seção para Anexar Documentos
    with st.expander("📎 Anexar Documentos", expanded=False):
        if not df_display.empty:
            idx_anexo = st.number_input(
                "Número da linha para anexar:",
                min_value=1,
                max_value=len(df_display),
                step=1,
                key="anexo_idx_pagar"
            )
            
            registro = df_display[df_display['#'] == idx_anexo].iloc[0]
            original_idx = df[df['fornecedor'] == registro['fornecedor']].index[0]
            
            uploaded_file = st.file_uploader(
                "Selecione o arquivo (PDF, JPG, PNG):",
                type=["pdf", "jpg", "png"],
                key=f"file_upload_pagar_{original_idx}"
            )
            
            if uploaded_file:
                destino = os.path.join(
                    ANEXOS_DIR,
                    "Contas a Pagar",
                    f"anexo_{aba}_{original_idx}_{uploaded_file.name}"
                )
                with open(destino, "wb") as f:
                    f.write(uploaded_file.getbuffer())
                st.success(f"Documento salvo em: {destino}")
                
                # Atualiza o registro com o caminho do anexo
                if uploaded_file.type == "application/pdf":
                    df.at[original_idx, 'boleto'] = destino
                else:
                    df.at[original_idx, 'comprovante'] = destino
                
                save_data(EXCEL_PAGAR, aba, df)

elif page == "Contas a Receber":
    st.subheader("🗂️ Contas a Receber")
    
    # Verificação do arquivo
    if not os.path.isfile(EXCEL_RECEBER):
        st.error(f"Arquivo '{EXCEL_RECEBER}' não encontrado. Verifique o caminho.")
        st.stop()
    
    # Seleção do mês
    existing_sheets = get_existing_sheets(EXCEL_RECEBER)
    aba = st.selectbox("Selecione o mês:", FULL_MONTHS, index=FULL_MONTHS.index(date.today().strftime("%B")))
    
    # Carrega os dados
    df = load_data(EXCEL_RECEBER, aba)
    
    # Filtro principal
    view_sel = st.radio("Visualizar:", ["Todos", "Recebidas", "Pendentes"], horizontal=True)
    
    if view_sel == "Recebidas":
        df_display = df[df["status_pagamento"] == "Recebido"].copy()
    elif view_sel == "Pendentes":
        df_display = df[df["status_pagamento"] != "Recebido"].copy()
    else:
        df_display = df.copy()
    
    # Adiciona numeração das linhas
    df_display.insert(0, '#', range(1, len(df_display) + 1))
    
    # Filtros avançados
    with st.expander("🔍 Filtros Avançados", expanded=False):
        col1, col2 = st.columns(2)
        with col1:
            cliente_filtro = st.selectbox(
                "Cliente",
                ["Todos"] + sorted(df["fornecedor"].dropna().unique().tolist()))
        with col2:
            status_filtro = st.selectbox(
                "Status",
                ["Todos"] + sorted(df["status_pagamento"].dropna().unique().tolist()))
    
    # Aplica filtros
    if cliente_filtro != "Todos":
        df_display = df_display[df_display["fornecedor"] == cliente_filtro]
    if status_filtro != "Todos":
        df_display = df_display[df_display["status_pagamento"] == status_filtro]
    
    # Exibe a tabela principal
    st.markdown("### 📋 Lançamentos")
    if df_display.empty:
        st.warning("Nenhum registro encontrado com os filtros selecionados.")
    else:
        # Seleciona e formata colunas para exibição
        cols_padrao = ['#', 'data_nf', 'fornecedor', 'valor', 'vencimento', 'status_pagamento', 'estado']
        cols_disponiveis = [c for c in cols_padrao if c in df_display.columns]
        
        df_exibicao = df_display[cols_disponiveis].copy()
        
        # Formatação
        if 'valor' in df_exibicao.columns:
            df_exibicao['valor'] = df_exibicao['valor'].apply(lambda x: f"R$ {float(x):,.2f}")
        if 'vencimento' in df_exibicao.columns:
            df_exibicao['vencimento'] = pd.to_datetime(df_exibicao['vencimento']).dt.strftime('%d/%m/%Y')
        if 'data_nf' in df_exibicao.columns:
            df_exibicao['data_nf'] = pd.to_datetime(df_exibicao['data_nf']).dt.strftime('%d/%m/%Y')
        
        st.dataframe(df_exibicao, height=400, use_container_width=True)
    
    # Seção de Edição
    with st.expander("✏️ Editar Registro", expanded=False):
        if not df_display.empty:
            idx_edicao = st.number_input(
                "Número da linha para editar:",
                min_value=1,
                max_value=len(df_display),
                step=1,
                key="edit_idx_receber"
            )
            
            registro = df_display[df_display['#'] == idx_edicao].iloc[0]
            original_idx = df[df['fornecedor'] == registro['fornecedor']].index[0]
            
            col1, col2 = st.columns(2)
            with col1:
                novo_valor = st.number_input(
                    "Valor (R$):",
                    value=float(registro['valor']),
                    step=0.01,
                    key="edit_valor_receber"
                )
                novo_vencimento = st.date_input(
                    "Vencimento:",
                    value=pd.to_datetime(registro['vencimento']).date(),
                    key="edit_venc_receber"
                )
            with col2:
                novo_estado = st.selectbox(
                    "Estado:",
                    options=["A Receber", "Recebido"],
                    index=0 if registro['estado'] == "A Receber" else 1,
                    key="edit_estado_receber"
                )
                nova_situacao = st.selectbox(
                    "Situação:",
                    options=["Em Atraso", "Recebido", "A Receber"],
                    index=0 if registro['situacao'] == "Em Atraso" else 1 if registro['situacao'] == "Recebido" else 2,
                    key="edit_situacao_receber"
                )
            
            if st.button("💾 Salvar Alterações", key="save_edit_receber"):
                df.at[original_idx, 'valor'] = novo_valor
                df.at[original_idx, 'vencimento'] = novo_vencimento
                df.at[original_idx, 'estado'] = novo_estado
                df.at[original_idx, 'situacao'] = nova_situacao
                
                if save_data(EXCEL_RECEBER, aba, df):
                    st.success("Registro atualizado com sucesso!")
                    st.experimental_rerun()
                else:
                    st.error("Erro ao salvar alterações.")
    
    # Seção para Remover Registros
    with st.expander("🗑️ Remover Registro", expanded=False):
        if not df_display.empty:
            idx_remocao = st.number_input(
                "Número da linha para remover:",
                min_value=1,
                max_value=len(df_display),
                step=1,
                key="remove_idx_receber"
            )
            
            registro = df_display[df_display['#'] == idx_remocao].iloc[0]
            original_idx = df[df['fornecedor'] == registro['fornecedor']].index[0]
            
            if st.button("Remover Registro", key="btn_remove_receber"):
                try:
                    df = df.drop(index=original_idx)
                    if save_data(EXCEL_RECEBER, aba, df):
                        st.success("Registro removido com sucesso!")
                        st.experimental_rerun()
                    else:
                        st.error("Erro ao salvar alterações.")
                except Exception as e:
                    st.error(f"Erro ao remover registro: {e}")
    
    # Seção para Adicionar Novos Registros
    with st.expander("➕ Adicionar Nova Conta", expanded=False):
        col1, col2 = st.columns(2)
        with col1:
            nova_data_nf = st.date_input("Data N/F:", value=date.today())
            nova_descricao = st.text_input("Descrição:")
            novo_cliente = st.text_input("Cliente:")
        with col2:
            novo_os = st.text_input("Documento/OS:")
            novo_vencimento = st.date_input("Vencimento:", value=date.today())
            novo_valor = st.number_input("Valor (R$):", min_value=0.01, step=0.01)
        
        novo_estado = st.selectbox("Estado:", ["A Receber", "Recebido"])
        nova_situacao = st.selectbox("Situação:", ["Em Atraso", "Recebido", "A Receber"])
        
        # Upload de arquivos
        col_anexo1, col_anexo2 = st.columns(2)
        with col_anexo1:
            boleto_file = st.file_uploader("Boleto (opcional):", type=["pdf", "jpg", "png"])
        with col_anexo2:
            comprovante_file = st.file_uploader("Comprovante (opcional):", type=["pdf", "jpg", "png"])
        
        if st.button("Adicionar Conta", key="btn_add_receber"):
            novo_registro = {
                "data_nf": nova_data_nf,
                "forma_pagamento": nova_descricao,
                "fornecedor": novo_cliente,
                "os": novo_os,
                "vencimento": novo_vencimento,
                "valor": novo_valor,
                "estado": novo_estado,
                "situacao": nova_situacao,
                "status_pagamento": "Recebido" if novo_estado == "Recebido" else "Pendente"
            }
            
            # Processa anexos
            if boleto_file:
                boleto_path = os.path.join(ANEXOS_DIR, "Contas a Receber", f"boleto_{uuid.uuid4()}.{boleto_file.name.split('.')[-1]}")
                with open(boleto_path, "wb") as f:
                    f.write(boleto_file.getbuffer())
                novo_registro["boleto"] = boleto_path
            
            if comprovante_file:
                comprovante_path = os.path.join(ANEXOS_DIR, "Contas a Receber", f"comprovante_{uuid.uuid4()}.{comprovante_file.name.split('.')[-1]}")
                with open(comprovante_path, "wb") as f:
                    f.write(comprovante_file.getbuffer())
                novo_registro["comprovante"] = comprovante_path
            
            if add_record(EXCEL_RECEBER, aba, novo_registro):
                st.success("Conta adicionada com sucesso!")
                st.experimental_rerun()
            else:
                st.error("Erro ao adicionar nova conta.")
    
    # Seção para Anexar Documentos
    with st.expander("📎 Anexar Documentos", expanded=False):
        if not df_display.empty:
            idx_anexo = st.number_input(
                "Número da linha para anexar:",
                min_value=1,
                max_value=len(df_display),
                step=1,
                key="anexo_idx_receber"
            )
            
            registro = df_display[df_display['#'] == idx_anexo].iloc[0]
            original_idx = df[df['fornecedor'] == registro['fornecedor']].index[0]
            
            uploaded_file = st.file_uploader(
                "Selecione o arquivo (PDF, JPG, PNG):",
                type=["pdf", "jpg", "png"],
                key=f"file_upload_receber_{original_idx}"
            )
            
            if uploaded_file:
                destino = os.path.join(
                    ANEXOS_DIR,
                    "Contas a Receber",
                    f"anexo_{aba}_{original_idx}_{uploaded_file.name}"
                )
                with open(destino, "wb") as f:
                    f.write(uploaded_file.getbuffer())
                st.success(f"Documento salvo em: {destino}")
                
                # Atualiza o registro com o caminho do anexo
                if uploaded_file.type == "application/pdf":
                    df.at[original_idx, 'boleto'] = destino
                else:
                    df.at[original_idx, 'comprovante'] = destino
                
                save_data(EXCEL_RECEBER, aba, df)
                
    st.markdown("---")
    st.subheader("💾 Exportar Aba Atual")
    try:
        df_to_save = load_data(EXCEL_RECEBER, aba)
        if not df_to_save.empty:
            save_data(EXCEL_RECEBER, aba, df_to_save)
        with open(EXCEL_RECEBER, "rb") as fx:
            st.download_button(
                label=f"Exportar '{aba}'",
                data=fx.read(),
                file_name=f"Contas a Receber - {aba}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    except Exception as e:
        st.error(f"Erro ao preparar download: {e}")

st.markdown("""
<div style="text-align: center; font-size:12px; color:gray; margin-top: 20px;">
    <p>© 2025 Desenvolvido por Vinicius Magalhães</p>
</div>
""", unsafe_allow_html=True)
